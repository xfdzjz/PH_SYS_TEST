
import time,sys
import openpyxl

testData = {}

# 可以用格式化字符串输出结果的用例
ptaFormatters = {
    'S2.1.1': [None, ["VBGS@3.3V","VBGS@2.2V","VBGS@5V"]],
    'S2.1.2': [None, ["VBGA@3.3V","VBGA@2.2V","VBGA@5V"]],
    'S2.1.3': [None, ["Freq@3.3V","Duty@3.3V","Freq@5V","Duty@5V","Freq@2.2V","Duty@2.2V",
        "Freq@3.3V","Duty@3.3V","Freq@5V","Duty@5V","Freq@2.2V","Duty@2.2V",
        "Freq@3.3V","Duty@3.3V","Freq@5V","Duty@5V","Freq@2.2V","Duty@2.2V",
        "Freq@3.3V","Duty@3.3V","Freq@5V","Duty@5V","Freq@2.2V","Duty@2.2V",
        "Freq@3.3V","Duty@3.3V","Freq@5V","Duty@5V","Freq@2.2V","Duty@2.2V",
    ]],
    'S2.1.4': [None, ["Freq@3.3V","Duty@3.3V","Freq@2.2V","Duty@2.2V","Freq@5V","Duty@5V"]],
    'S2.1.5': [None, ["Freq@3.3V","Duty@3.3V","Freq@2.2V","Duty@2.2V","Freq@5V","Duty@5V"]],
    'S2.1.6': [None, ["Vref@1.2V,VCC3.3","Vref@1.2V,VCC5","Vref@1.2V,VCC2.2",
        "Vref@1.5V,VCC3.3","Vref@1.5V,VCC5","Vref@1.5V,VCC2.2",
        "Vref@2.0V,VCC3.3","Vref@2.0V,VCC5","Vref@2.0V,VCC2.2",
        "Vref@2.5V,VCC3.3","Vref@2.5V,VCC5","Vref@2.5V,VCC2.2"
    ]],
    'S2.2.1': [None, ["V1P5S@1.5V,VCC=3.3V","V1P5S@1.5V,VCC=5V","V1P5S@1.5V,VCC=2.2V",
        "V1P5S@1.2V,VCC=3.3V","V1P5S@1.2V,VCC=5V","V1P5S@1.2V,VCC=2.2V",
    ]],
    'S2.2.3': [None, ["V1P5S_L"]],
    'S2.2.4': [None, ["V1P5D@1.5V,VCC=3.3V","V1P5D@1.5V,VCC=5V","V1P5D@1.5V,VCC=2.2V",
        "V1P5D@1.2V,VCC=3.3V","V1P5D@1.2V,VCC=5V","V1P5D@1.2V,VCC=2.2V",
    ]],
    'S2.2.6': [None, ["V1P5D_L"]],
    'S2.2.7': [None, ["V1P5A@3.3V","V1P5A@5V","V1P5A@2.2V",
        "V1P5A@3.3V","V1P5A@5V","V1P5A@2.2V",
        "V1P5A@3.3V","V1P5A@5V","V1P5A@2.2V",
        "V1P5A@3.3V","V1P5A@3.3V",
    ]],
    # 'S2.2.8': 'ready:(\S+)mv VCC is (\S+)v avg is (\S+)',
    # 'S2.2.9': 'ready:(\S+)mv VCC is (\S+)v avg is (\S+)',
    'S2.2.10': [None, ["Visrch"]],
    'S2.2.11': [None, ["Visrca"]],
    'S2.3': [None, ["Isrcs@3V","Isrcs@5V","Isrcs@2.2V"]],
    'S2.4.1': [None, ["Vhy@0","Vhy@12","Vhy@25","Vhy@50"]],
    # 'S2.4.2': 'NOTHING',  # nothinf to parse
    'S2.5': [None, [], False],
    # 'S2.6.1': 'indled amp is (\S+) when VCC is (\S+)v',
    # 'S2.7.1': 'ready:(\S+)mv VCC is (\S+)v avg is (\S+)',  # not tested yet
    # 'S2.7.2': 'ready:(\S+)mv VCC is (\S+)v avg is (\S+)',  # not tested yet
    # 'S2.8.1': 'resp is\s+(\S+).*?apply (\S+)',  # not tested yet
    # 'S2.8.2': 'ready:(\S+)mv VCC is (\S+)v avg is (\S+)',  # not tested yet
    # 'S2.8.3': 'ready:(\S+)mv VCC is (\S+)v avg is (\S+)',  # not tested yet
    # 'S2.8.4': 'ready:(\S+)mv VCC is (\S+)v avg is (\S+)',  # not tested yet
    'S3.1.1': [None, ["Ivcc@5V","Ivcc@3.3V","Ivcc@2.2V"]],
    'S3.1.2': [None, ["Idsleep@5V","Idsleep@3.3V","Idsleep@2.2V"]],
    'S3.1.3': [None, ["Ipwrdown@5V","Ipwrdown@3.3V","Ipwrdown@2.2V"]],
    'S3.1.4': [None, ["I_BVS@3.3V"]],
    'S3.1.5': [None, ["Isleep@5V","Isleep@3.3V","Isleep@2.2V"]],
    # 'S3.2.1': [[0, 3], ["VTR", "VTF"]],
    # 'S3.2.2': [[0, 1], ["Tfpor", "Tspor"]],
    'S3.3.1': [[0,2,4],["Itherm@3.3V","Itherm@2.2V","Itherm@5V"]],
    # 'S3.3.2': [[0,2,4],["Itherm@3.3V","Itherm@2.2V","Itherm@5V"]],
    'S3.3.3': [[0,2],["Ibled@3.3V","Irbled@3.3V"]],
    'S3.4': [None,["Freq1","Duty1","Freq2","Duty2"]],

    # 'S4.1': 'NOTHING',  # nothinf to parse
    "S4.2": [[2,5,8,11,14,17,20,23,26,29],["IN1@0.5","IN1@1.5","IN5@0.5","IN5@1.5",
        "IN6@0.5","IN6@1.5","IN8@0.5","IN8@1.5","IN9@0.5","IN9@1.5"
    ], False],

    "S5": [[2,5,8,11,14,17,20,23,26,29,32,35,38,41],["IN0@0.5","IN0@1.5","IN2@0.5","IN2@1.5","IN3@0.5","IN3@1.5",
        "IN4@0.5","IN4@1.5","IN7@0.5","IN7@1.5","IN10@0.5","IN10@1.5","IN11@0.5","IN11@1.5"
    ], False],

    "S6.1": [None,["Duty1","Freq1","Duty2","Freq2","Duty3","Freq3","Duty4","Freq4",
        "Duty5","Freq5","Duty6","Freq6","Duty7","Freq7","Duty8","Freq8"
    ]],
    "S6.2": [None,["VOK1","VOK2","VOK3","VOK4","VOK1","VOK2","VOK3","VOK4",
        "VOK1","VOK2","VOK3","VOK4","VOK1","VOK2","VOK3","VOK4"
    ]],
    "S6.3": [None,["LX","VH","LX","VH","LX","VH","LX","VH"]],

}


def printWithFormatter(test_case, ts, sn, result, data):
    exectime = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(ts))
    print("用例:%s, 执行时间:%s, SN=%s, 执行结果: %s" %
          (test_case, exectime, sn, result))
    params = []  # 输出参数， 数组
    allParams = []
    for t in data:  # data为元组数组, 展开成一维数组allParams
        if isinstance(t,tuple):
            t = list(t)
            allParams.extend(t)
        elif isinstance(t,list):
            allParams.extend(t)
        else:
            allParams.append(t)

    if test_case in ptaFormatters:
        fmtFilter = ptaFormatters[test_case][0]  # 参数过滤器
        if fmtFilter == None: # 不需要过滤
            params = allParams
        else:
            for i in fmtFilter:
                params.append(allParams[i])
    else:
        params = allParams

    if not test_case in testData:
        testData[test_case] = {sn: {"time": exectime, "result": result, "data": params}}
    else:
        testData[test_case][sn] = {"time": exectime, "result": result, "data": params}
    

# 特殊的输出函数，函数名为printCase_Name
# CaseName经过.=>_变换

def printAll(resultExcel):
    workbook = openpyxl.Workbook()
    sheetIndex = 0
    for case in testData:
        print("export test result for %s:" % case)
        sheet = workbook.create_sheet(index=sheetIndex, title=case)
        sheetIndex = sheetIndex + 1
        results = testData[case]
        sheet["A1"] = "执行时间"
        sheet["B1"] = "序列号"
        sheet["C1"] = "执行结果"
        if case in ptaFormatters:
            fields = ptaFormatters[case][1]
            for i in range(len(fields)):
                sheet.cell(row=1,column=4+i).value = fields[i]
        row = 2
        for sn in results:
            result = results[sn]
            sheet["A%d" % (row)] = result["time"]
            sheet["B%d" % (row)] = sn
            sheet["C%d" % (row)] = result["result"]
            for i in range(len(result["data"])):
                value = None
                if case in ptaFormatters:
                    if len(ptaFormatters[case]) > 2:
                        value = result["data"][i]
                if value == None:
                    try:
                        value = float(result["data"][i])
                    except:
                        value = result["data"][i]
                sheet.cell(row=row,column=4+i).value = value
            row = row + 1

    workbook.save(resultExcel)

if __name__ == "__main__":
    pass
